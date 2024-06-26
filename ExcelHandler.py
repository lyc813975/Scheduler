import os
import openpyxl
import openpyxl.cell
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from Util import *

class ExcelHandler:
    def __init__(
        self,
        filename = "sheet.xlsx"
    ):
        self.workbook = None
        self.path = os.getcwd()
        self.filename = filename

    def load(self):
        if self.workbook is not None:
            print("[Warning]: Discard Current Workbook")
        self.workbook = openpyxl.load_workbook(os.path.join(self.path, self.filename))

    def create_worksheet(self):
        if self.workbook is not None:
            print("[Warning]: Discard Current Workbook")
        self.workbook = openpyxl.Workbook()

    def create_sheet(self, title=None, pos=None):
        self.workbook.create_sheet(title, pos)

    def save_worksheet(self, path=None):
        if path is not None:
            self.path, self.filename = os.path.split(path)
        self.workbook.save(os.path.join(self.path, self.filename))

    def get_filename(self):
        return self.filename

    def get_filepath(self):
        return self.path

    def get_sheet_titles(self):
        return self.workbook.sheetnames

    def get_sheet_max_row(self, title):
        return self.workbook[title].max_row

    def get_sheet_max_column(self, title):
        return self.workbook[title].max_column

    def unmerge_cells(self, title, start_row, start_col, end_row, end_col):
        raise NotImplementedError

    def merge_cells(self, title, start_row, start_col, end_row, end_col):
        if isinstance(start_row, int):
            start_row = convert_python_row_to_excel_row(start_row)
        if isinstance(end_row, int):
            end_row = convert_python_row_to_excel_row(end_row)
        if isinstance(start_col, int):
            start_col = convert_python_column_to_excel_column(start_col)
        if isinstance(end_col, int):
            end_col = convert_python_column_to_excel_column(end_col)
        self.workbook[title].merge_cells(f"{start_col}{start_row}:{end_col}{end_row}")

    def get_cell(self, title, row, col, return_root=True, return_range=False):
        if isinstance(row, int):
            row = convert_python_row_to_excel_row(row)
        if isinstance(col, int):
            col = convert_python_column_to_excel_column(col)
        try:
            cell = self.workbook[title][col+row]
            merged = None
            if return_root and isinstance(cell, openpyxl.cell.cell.MergedCell):
                merged = root_of_merged_cell(self.workbook[title], col+row)
                cell = self.workbook[title][merged.start_cell.coordinate]
            if return_range:
                return cell, merged
            return cell
        except ValueError as e:
            print(e)

    def set_cell(self, title, row, col, cell):
        if isinstance(row, int):
            row = convert_python_row_to_excel_row(row)
        if isinstance(col, int):
            col = convert_python_column_to_excel_column(col)
        try:
            self.workbook[title][col+row] = cell
        except ValueError as e:
            print(e)

    def get_cell_value(self, title, row, col):
        return self.get_cell(title, row, col).value

    def set_cell_value(self, title, row, col, value):
        self.get_cell(title, row, col).value = value

    def get_row_value(self, title, row):
        max_column = self.get_sheet_max_column(title)
        values = []
        for i in range(max_column):
            value = self.get_cell(title, row, i).value
            values.append(value)
        return values

    def get_col_value(self, title, col):
        max_row = self.get_sheet_max_row(title)
        values = []
        for i in range(max_row):
            values.append(self.get_cell(title, i, col).value)
        return values
    
    def get_cell_font(self, title, row, col):
        return self.get_cell(title, row, col).font

    def set_cell_font(self, title, row, col, font):
        self.get_cell(title, row, col).font = font

    def get_cell_fill(self, title, row, col):
        return self.get_cell(title, row, col).fill

    def set_cell_fill(self, title, row, col, fill):
        self.get_cell(title, row, col).fill = fill

    def get_cell_border(self, title, row, col):
        return self.get_cell(title, row, col).border

    def set_cell_border(self, title, row, col, border):
        cell, cellRange = self.get_cell(title, row, col, return_range=True)
        if cellRange is None:
            cell.border = border
        else:
            for i in range(cellRange.min_row - 1, cellRange.max_row):
                for j in range(cellRange.min_col - 1, cellRange.max_col):
                    self.get_cell(title, i, j, return_root=False).border = border

    def get_column_width(self, title, col):
        if isinstance(col, int):
            col = convert_python_column_to_excel_column(col + 1)
        return self.workbook[title].column_dimensions[col].width

    def set_column_width(self, title, col, width):
        if isinstance(col, int):
            col = convert_python_column_to_excel_column(col + 1)
        self.workbook[title].column_dimensions[col].width = width

    def get_row_height(self, title, row):
        if isinstance(row, int):
            row = convert_python_row_to_excel_row(row)
        if isinstance(row, str):
            row = int(row)
        return self.workbook[title].row_dimensions[row].height

    def set_row_height(self, title, row, height):
        if isinstance(row, int):
            row = convert_python_row_to_excel_row(row)
        if isinstance(row, str):
            row = int(row)
        self.workbook[title].row_dimensions[row].height = height


def test_writer():
    writer = ExcelHandler()
    writer.create_worksheet()
    writer.create_sheet()
    titles = writer.get_sheet_titles()

    writer.set_cell_value(titles[0], 0, 0, 1)
    writer.set_cell_font(titles[0], 0, 0, Font("新細明體", sz=12, bold=False, italic=False, underline=None, color='00000000'))

    writer.set_cell_value(titles[0], 0, 1, 2)
    writer.set_cell_font(titles[0], 0, 1, Font("新細明體", sz=12, bold=False, italic=False, underline=None, color='FFFF0000'))
    
    writer.set_cell_value(titles[0], 0, 2, 3)
    writer.set_cell_font(titles[0], 0, 2, Font("新細明體", sz=12, bold=False, italic=False, underline=None, color='FFFFC000'))
    writer.set_cell_fill(titles[0], 0, 2, PatternFill(fill_type="solid", fgColor="FF46A2D5"))

    writer.set_cell_value(titles[0], 1, 0, "a")
    writer.set_cell_font(titles[0], 1, 0, Font("新細明體", sz=12, bold=True, italic=False, underline=None, color='00000000'))

    writer.set_cell_value(titles[0], 1, 1, "b")
    writer.set_cell_font(titles[0], 1, 1, Font("新細明體", sz=12, bold=False, italic=True, underline=None, color='00000000'))

    writer.set_cell_value(titles[0], 1, 2, "c")
    writer.set_cell_font(titles[0], 1, 2, Font("新細明體", sz=12, bold=False, italic=False, underline='single', color='00000000'))

    writer.set_row_height(titles[1], 0, 99)
    writer.set_column_width(titles[1], 0, 6.5)
    writer.set_cell_value(titles[1], 0, 0, "你")
    writer.set_cell_font(titles[1], 0, 0, Font("新細明體", sz=12, bold=False, italic=False, underline=None, color='00000000'))

    writer.set_column_width(titles[1], 1, 8)
    writer.set_cell_value(titles[1], 0, 1, "好")
    writer.set_cell_font(titles[1], 0, 1, Font("新細明體", sz=16, bold=False, italic=False, underline=None, color='00000000'))
    
    writer.set_column_width(titles[1], 2, 13)
    writer.set_cell_value(titles[1], 0, 2, "啊")
    writer.set_cell_font(titles[1], 0, 2, Font("新細明體", sz=24, bold=False, italic=False, underline=None, color='00000000'))

    writer.merge_cells(titles[1], 1, 1, 3, 2)
    writer.set_cell_value(titles[1], 1, 2, "啊")
    writer.set_cell_border(titles[1], 2, 2, Border(left=Side(style='thin')))

    writer.save_worksheet()


def test_reader():
    reader = ExcelHandler()
    reader.load()
    titles = reader.get_sheet_titles()
    file = reader.get_filename()
    
    print(f"Excel Name: {file}")
    for title in titles:
        for i in range(reader.get_sheet_max_row(title)):
            for j in range(reader.get_sheet_max_column(title)):
                value = reader.get_cell_value(title, i, j)
                if value is None:
                    continue
                fill = reader.get_cell_fill(title, i, j)
                font = reader.get_cell_font(title, i, j)
                border = reader.get_cell_border(title, i, j)
                color = font.color
                print(f"Title: {title}, Cell Position: {i, j}")
                print(f"\tValue: {value}, Height: {reader.get_row_height(title, i)}, Width: {reader.get_column_width(title, j)}")
                print(f"\tFont: {font.name}, Size: {font.size}, Bold: {font.bold}, Italic: {font.italic}, Underline: {font.underline}")
                print(f"\tText Color: {get_color_value(color)}, Background color: {get_color_value(fill.fgColor)}")
                print(f"\tLeft: {border.left.style if border.left is not None else border.left}, ", end="")
                print(f"Right: {border.right.style if border.right is not None else border.right}, ", end="")
                print(f"Top: {border.top.style if border.top is not None else border.top}, ", end="")
                print(f"Bottom: {border.bottom.style if border.bottom is not None else border.bottom}")

if __name__ == "__main__":
    test_writer()
    test_reader()